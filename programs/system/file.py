import json
import re
from os import listdir
from os.path import isdir
from typing import Union, TextIO
from openpyxl import load_workbook
import PyPDF2
from zipfile import ZipFile


def take_desktop() -> str:
    return r"C:\Users\maxma\Desktop"


def file_from_desktop(file: str) -> str:
    return rf"C:\Users\maxma\Desktop\{file}"


def split_name_format(name_format,
                      mode='l',
                      prefix='@'):
    '''

    name_format: Имя файла с расширением

    mode: Что вернуть. Строка без пробелов

    n: Имя,

    f: Формат,

    l: Список: Имя и формат,

    p: Префикс


    d: Добавить к формату точку
    '''

    name_format_list = name_format.split('.')
    form = name_format_list[-1]
    name = '.'.join(name_format_list[:-1])

    if 'd' in mode:
        form = '.' + form

    if 'l' in mode:
        return [name, form]
    if 'f' in mode:
        return form
    if 'n' in mode:
        return name
    if 'p' in mode:
        return name.split(prefix)[-1]


def make_json(path2file: str, content: Union[list, dict]):
    with open(path2file, 'w+', encoding="utf-8") as file:
        json.dump(content, file, ensure_ascii=False)


def read_json(path2file: str) -> Union[list, dict]:
    with open(path2file, 'r', encoding='utf-8') as file:
        return json.load(file)


def make_file(path2file: str) -> TextIO:
    file = open(path2file, 'w+', encoding="utf-8")
    return file


def make_svg(path2file: str, width: Union[int, float] = 1000, height: Union[int, float] = 1000) -> TextIO:
    file = open(path2file, 'w+', encoding="utf-8")
    file.write(f'<?xml version="1.0" encoding="utf-8"?>\n<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}">')
    return file


def read_file(path2file: str) -> str:
    with open(path2file, 'r', encoding='utf-8') as file:
        return file.read()


def make_zip(path2zip, path2file, path2file_in_zip=None):
    realpath = path2file if path2file_in_zip is None else path2file_in_zip

    with ZipFile(path2zip, 'a') as dir_zip:
        dir_zip.write(path2file, arcname=realpath)


def read_pdf(path2file: str, startpage: int = 0, stoppage: int = -1) -> str:
    text = ""
    with open(path2file, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        if stoppage <= 0:
            stoppage = len(reader.pages)

        for i in range(startpage, stoppage):
            text += reader.pages[i].extract_text() + "\n"

    return text


def read_xls(path2file: str) -> list:
    workbook = load_workbook(path2file)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    return data


def read_shortcut(path2file: str) -> str:
    pattern = r"URL=(.*)"
    with open(path2file, 'r', encoding='utf-8') as file:
        data = file.read()

    match = re.search(pattern, data)
    if match:
        return match.group(1).strip()
    return ''


if __name__ == "__main__":
    pass
