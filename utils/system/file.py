import json
import re
from os import listdir
from os.path import isdir
from typing import Union, TextIO
from openpyxl import load_workbook
import PyPDF2
from zipfile import ZipFile
from utils.config import WINDOWS_USERNAME


def take_desktop() -> str:
    return rf"C:\Users\{WINDOWS_USERNAME}\Desktop"


def file_from_desktop(file: str) -> str:
    return rf"C:\Users\{WINDOWS_USERNAME}\Desktop\{file}"


def split_filename(filename: str, mode: str = 'nf', prefix: str = '@'):
    parts = filename.split('.')
    if len(parts) == 1:
        return parts[0]
    extension = parts[-1].lower()
    name = '.'.join(parts[:-1])

    if 'd' in mode:
        extension = '.' + extension

    output = []

    if 'n' in mode:
        output.append(name)
    if 'f' in mode:
        output.append(extension)
    if 'p' in mode:
        output.append(name.split(prefix)[-1])

    return output[0] if len(output) == 1 else output


def make_json(path2file: str, content: Union[list, dict]):
    with open(path2file, 'w+', encoding="utf-8") as file:
        json.dump(content, file, ensure_ascii=False)


def read_json(path2file: str) -> Union[list, dict]:
    with open(path2file, 'r', encoding='utf-8') as file:
        return json.load(file)


def make_file(path2file: str) -> TextIO:
    file = open(path2file, 'w+', encoding="utf-8")
    return file


def read_file(path2file: str) -> str:
    with open(path2file, 'r', encoding='utf-8') as file:
        return file.read()


def make_zip(path2zip, path2file, path2file_in_zip=None):
    realpath = path2file if path2file_in_zip is None else path2file_in_zip

    with ZipFile(path2zip, 'a') as dir_zip:
        dir_zip.write(path2file, arcname=realpath)


def read_pdf(path2file: str, startpage: int = 0, stoppage: int = -1) -> str:
    text = ""
    with open(path2file, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        if stoppage <= 0:
            stoppage = len(reader.pages)

        for i in range(startpage, stoppage):
            text += reader.pages[i].extract_text() + "\n"

    return text


def read_xls(path2file: str) -> list:
    workbook = load_workbook(path2file)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    return data


def read_shortcut(path2file: str) -> str:
    pattern = r"URL=(.*)"
    with open(path2file, 'r', encoding='utf-8') as file:
        data = file.read()

    match = re.search(pattern, data)
    if match:
        return match.group(1).strip()
    return ''


if __name__ == "__main__":
    pass
